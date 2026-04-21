#!/usr/bin/env python3
"""Build roster-valid.xlsx from CSV and verify all fixtures are well-formed."""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent
CSV_PATH = ROOT / "roster-valid.csv"
JSON_PATH = ROOT / "roster-valid.json"
XLSX_PATH = ROOT / "roster-valid.xlsx"

MAX_PER_SIDE = 20


def normalize_type(raw: str) -> str:
    s = (raw or "").strip().lower()
    if s in ("corp", "corporate", "corporates"):
        return "Corp"
    if s in ("startup", "startups", "su"):
        return "Startup"
    return ""


def validate_participants(rows: list[dict]) -> list[str]:
    """Rules aligned with validateRoster / assignIdsFromRows in index.html."""
    errors: list[str] = []
    if not rows:
        errors.append("Roster is empty.")
        return errors
    seen: set[str] = set()
    n_corp = n_su = 0
    for p in rows:
        pid = str(p.get("id") or "").strip()
        if not pid:
            errors.append("Every row needs a non-empty id after normalization.")
            continue
        if pid in seen:
            errors.append(f"Duplicate id: {pid}")
        seen.add(pid)
        name = str(p.get("name") or "").strip()
        sector = str(p.get("sector") or "").strip()
        typ = p.get("type")
        if not name:
            errors.append(f"Empty name for id {pid}")
        if not sector:
            errors.append(f"Empty sector for {name or pid}")
        if typ not in ("Corp", "Startup"):
            errors.append(f"Invalid type for {name or pid}")
            continue
        if typ == "Corp":
            n_corp += 1
        else:
            n_su += 1
    if n_corp < 1:
        errors.append("Need at least one corporate.")
    if n_su < 1:
        errors.append("Need at least one startup.")
    if n_corp > MAX_PER_SIDE:
        errors.append(f"Too many corporates (max {MAX_PER_SIDE}).")
    if n_su > MAX_PER_SIDE:
        errors.append(f"Too many startups (max {MAX_PER_SIDE}).")
    return errors


def assign_ids_from_rows(raw_rows: list[dict]) -> list[dict]:
    """Mirror index.html assignIdsFromRows + normalizeType."""
    n_c = 0
    n_s = 0
    out: list[dict] = []
    used: set[str] = set()

    for raw in raw_rows:
        typ = normalize_type(str(raw.get("type", "")))
        if not typ:
            continue
        raw_id = raw.get("id")
        pid = str(raw_id).strip() if isinstance(raw_id, str) and raw_id.strip() else ""
        if not pid:
            pid = ("C" if typ == "Corp" else "S") + str(n_c + 1 if typ == "Corp" else n_s + 1)
            if typ == "Corp":
                n_c += 1
            else:
                n_s += 1
            while pid in used or any(x["id"] == pid for x in out):
                if typ == "Corp":
                    n_c += 1
                    pid = "C" + str(n_c)
                else:
                    n_s += 1
                    pid = "S" + str(n_s)
        used.add(pid)
        out.append(
            {
                "id": pid,
                "name": str(raw.get("name") or "").strip(),
                "sector": str(raw.get("sector") or "").strip(),
                "type": typ,
            }
        )
    return out


def load_csv_rows() -> list[dict]:
    text = CSV_PATH.read_text(encoding="utf-8").replace("\ufeff", "")
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    header = [h.strip().lower() for h in lines[0].split(",")]

    def ix(name: str) -> int:
        return header.index(name)

    i_name, i_type, i_sector = ix("name"), ix("type"), ix("sector")
    raw = []
    for ln in lines[1:]:
        cells = [c.strip().strip('"') for c in ln.split(",")]
        raw.append(
            {"name": cells[i_name], "type": cells[i_type], "sector": cells[i_sector]}
        )
    return raw


def load_json_rows() -> list[dict]:
    data = json.loads(JSON_PATH.read_text(encoding="utf-8"))
    arr = data if isinstance(data, list) else data["participants"]
    return [
        {
            "id": row.get("id"),
            "name": row.get("name"),
            "type": row.get("type"),
            "sector": row.get("sector"),
        }
        for row in arr
    ]


def load_xlsx_rows() -> list[dict]:
    from openpyxl import load_workbook

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header_row = next(it, None)
    if not header_row:
        raise RuntimeError("Empty spreadsheet")
    keys = [str(c).strip() if c is not None else "" for c in header_row]

    def find_col(want: str) -> int:
        w = want.lower()
        for i, k in enumerate(keys):
            if k.strip().lower() == w:
                return i
        return -1

    i_name = find_col("name")
    i_type = find_col("type")
    i_sector = find_col("sector")
    i_id = find_col("id")
    if i_name < 0 or i_type < 0 or i_sector < 0:
        raise RuntimeError("Spreadsheet first sheet must include columns: name, type, sector")

    raw = []
    for row in it:
        if row is None:
            continue
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        def cell(idx: int) -> str:
            if idx < 0 or idx >= len(row):
                return ""
            v = row[idx]
            return "" if v is None else str(v).strip()

        item = {
            "name": cell(i_name),
            "type": cell(i_type),
            "sector": cell(i_sector),
        }
        if i_id >= 0:
            cid = cell(i_id)
            if cid:
                item["id"] = cid
        raw.append(item)
    return raw


def write_xlsx() -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Roster"
    ws.append(["name", "type", "sector"])
    for line in CSV_PATH.read_text(encoding="utf-8").splitlines()[1:]:
        line = line.strip()
        if not line:
            continue
        ws.append([c.strip().strip('"') for c in line.split(",")])
    wb.save(XLSX_PATH)


def main() -> int:
    write_xlsx()
    print(f"Wrote {XLSX_PATH}")

    results: dict[str, list[dict]] = {}
    for label, raw_loader in (
        ("CSV", load_csv_rows),
        ("JSON", load_json_rows),
        ("XLSX", load_xlsx_rows),
    ):
        try:
            raw = raw_loader()
            participants = assign_ids_from_rows(raw)
        except Exception as e:
            print(f"FAIL {label}: {e}")
            return 1
        err = validate_participants(participants)
        if err:
            print(f"FAIL {label}: " + " ".join(err))
            return 1
        results[label] = participants
        print(f"OK   {label}: {len(participants)} participants")

    # Same people (name+type+sector), ids may differ for auto-assigned CSV vs explicit JSON
    def keyset(participants: list[dict]) -> set[tuple[str, str, str]]:
        return {(p["name"], p["type"], p["sector"]) for p in participants}

    k_csv = keyset(results["CSV"])
    if k_csv != keyset(results["JSON"]) or k_csv != keyset(results["XLSX"]):
        print("FAIL: name/type/sector sets differ between CSV, JSON, XLSX")
        return 1

    print("All fixture checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
